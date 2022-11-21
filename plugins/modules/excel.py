#!/usr/bin/python
# -*- coding: utf-8 -*-


# Copyright: (c) 2022, DELGEHIER Cedric <cedric.delgehier@gmail.com>
# GNU General Public License v3.0+ (see COPYING or https://www.gnu.org/licenses/gpl-3.0.txt)

from ansible.module_utils.basic import AnsibleModule

# from ansible.module_utils.basic import env_fallback
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import os

__metaclass__ = type


ANSIBLE_METADATA = {
    "metadata_version": "1.0",
    "status": ["preview"],
    "supported_by": "community",
}


DOCUMENTATION = """
---
module: excel
short_description: Write data to excel spreadsheet.
description:
  - Write data to excel spreadsheet.
  - The data needs to be a list of dictonaries.
  - The keys of the first dictionnary are used as headers for the excel spreadsheet.

version_added: "2.9"

options:

  column_width:
    type: str
    description:
      - Width of columns for the sheet.
      - It can be a size integer or the string "auto".
      - If this size starts width a "<" sympbol, the width is in the autofit mode but limited to the interger specified.
    required: false
    default: auto

  create:
    type: bool
    description:
      - Force creation when file exists
    required: false
    default: false

  data:
    type: list
    description:
      - The data to write in the file.
      - Data mmust be a list of dicts.
    required: false

  file:
    type: path
    description:
      - The file name.
    required: true
    aliases: [ workbook ]

  operation:
    type: str
    description:
      - The operation to do (read or write).
    required: true
    choices:
      - write

  path:
    type: path
    description:
      - the path containing the file.
    required: true

  table_name:
    type: str
    description:
      - Name of the table that will contain data.
    required: false

  worksheet:
    type: str
    description:
      - The name of worksheet to proceed.
    required: true



author: "Cédric DELGEHIER (@cdelgehier)"
requirements:
  - openpyxl
  - os
"""

EXAMPLES = """
- name: "Add a sheet in a non existant workbook"
    excel:
    operation: write
    data: "{{ data }}"
    table_name: pop_first_names
    path: "{{ playbook_dir }}"
    file: "test1.xlsx"
    worksheet: "my_title"
    create: true
    column_width: "<42"
    #column_width: 50
"""


def main():

    argument_spec = dict(
        column_width=dict(type="str", required=False, default="auto"),
        create=dict(type="bool", required=False, default=False),
        data=dict(type="list", required=False),
        file=dict(type="str", required=True, aliases=["workbook"]),
        operation=dict(
            type="str",
            default="write",
            choices=[
                "write",
                # "read",
            ],
        ),
        path=dict(type="path", required=True),
        table_name=dict(type="str", required=False),
        worksheet=dict(type="str", required=True),
    )

    module = AnsibleModule(argument_spec=argument_spec, supports_check_mode=False)

    # Extract our parameters
    column_width = module.params.get("column_width")
    create = module.params.get("create")
    data = module.params.get("data")
    operation = module.params.get("operation")
    if data is None and operation == "write":
        module.fail_json(
            msg="The data paramater can't be None with the operation 'write'."
        )
    file = module.params.get("file")
    if not file.endswith(".xlsx"):
        module.fail_json(
            msg="openpyxl does not support file format, only xlsx is supported for this module",
        )

    path = module.params.get("path")
    table_name = module.params.get("table_name")
    worksheet = module.params.get("worksheet")

    if operation == "write":
        if not os.path.exists(path):
            if not create:
                module.fail_json(
                    msg="The path {} doesn't exist and the parameter 'create' is false.".format(
                        path
                    )
                )
            else:
                os.mkdir(path)

        file_fullpath = os.path.join(path, file)
        if not os.path.isfile(file_fullpath):
            # file doesn't exist
            if not create:
                module.fail_json(
                    msg="The file {} doesn't exist and the parameter 'create' is false.".format(
                        file_fullpath
                    )
                )
            else:
                workbook = Workbook()
        else:

            workbook = openpyxl.load_workbook(filename=file_fullpath, data_only=True)

        sheetnames = workbook.sheetnames
        if worksheet in sheetnames:
            # sheet already exists
            workbook.remove(workbook[worksheet])

        new_worksheet = workbook.create_sheet(title=worksheet)
        if "Sheet" in sheetnames:
            # sheet already exists
            workbook.remove(workbook["Sheet"])


        # write data
        headers = list(data[0].keys())
        new_worksheet.append(headers)
        for line in data:
            new_worksheet.append(list(line.values()))

        # create a table
        if table_name is not None:

            # data_range = "A1:" + get_column_letter(new_worksheet.max_column) + str(new_worksheet.max_row)
            data_range = new_worksheet.calculate_dimension()
            mediumStyle = openpyxl.worksheet.table.TableStyleInfo(
                name="TableStyleMedium9", showRowStripes=True
            )
            table = Table(
                displayName=table_name, ref=data_range, tableStyleInfo=mediumStyle
            )

            new_worksheet.add_table(table)

        # Adjust column width
        for column in new_worksheet.columns:
            letter = column[0].column_letter # Get the column name

            if column_width.isnumeric():
                # fixed
                new_worksheet.column_dimensions[letter].width = column_width
            else:
                # auto
                max_length = 0
                letter = column[0].column_letter # Get the column name
                for cell in column:
                    try: # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = int((max_length + 2) * 1.2)

                if column_width.startswith('<'):
                    column_width_int = int(column_width.split('<')[1])
                    if adjusted_width > column_width_int:
                        adjusted_width = column_width_int

                new_worksheet.column_dimensions[letter].width = adjusted_width


        workbook.save(file_fullpath)
        module.exit_json(
            changed=True,
        )


if __name__ == "__main__":
    main()
